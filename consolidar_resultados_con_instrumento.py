from __future__ import annotations

import argparse
import html
import re
import unicodedata
from pathlib import Path
from typing import Callable, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

VALID_ANSWER_CODES = {"A", "B", "C", "D", "E", "N", "RD"}
HEADER_SCAN_REQUIRED_MIN_P = 5

DEFAULT_BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_DIR = DEFAULT_BASE_DIR / "inputs"
DEFAULT_OUTPUT_DIR = DEFAULT_BASE_DIR / "outputs"

PAES_M2_2024_TABLE = {
    0: 100,
    1: 163,
    2: 195,
    3: 223,
    4: 248,
    5: 271,
    6: 292,
    7: 312,
    8: 330,
    9: 348,
    10: 366,
    11: 381,
    12: 395,
    13: 408,
    14: 422,
    15: 437,
    16: 452,
    17: 466,
    18: 478,
    19: 488,
    20: 498,
    21: 509,
    22: 522,
    23: 537,
    24: 551,
    25: 563,
    26: 573,
    27: 582,
    28: 592,
    29: 603,
    30: 616,
    31: 631,
    32: 645,
    33: 657,
    34: 667,
    35: 678,
    36: 689,
    37: 703,
    38: 718,
    39: 734,
    40: 748,
    41: 762,
    42: 776,
    43: 792,
    44: 810,
    45: 829,
    46: 848,
    47: 868,
    48: 890,
    49: 915,
    50: 1000,
}


def strip_accents(text: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c))


def normalize_col_name(value: object) -> str:
    if value is None:
        return ""
    text = strip_accents(str(value)).replace("\xa0", " ").replace("\n", " ").strip().lower()
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^a-z0-9_]+", "", text)
    return text.strip("_")


def slug_label(value: object) -> str:
    text = normalize_col_name(value)
    return re.sub(r"_+", "_", text)


def normalize_rut(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = strip_accents(str(value)).upper()
    text = re.sub(r"[^0-9K]", "", text)
    if not text:
        return ""
    if len(text) == 1:
        return text
    return f"{text[:-1]}-{text[-1]}"


def normalize_yes(value: object) -> str:
    text = strip_accents("" if value is None else str(value)).strip().lower()
    if not text:
        return ""
    if text.startswith("si") or "autorizo" in text or text in {"yes", "y", "true", "1"}:
        return "SI"
    if text.startswith("no") or text in {"false", "0"}:
        return "NO"
    return text.upper()


def try_read_with_header(path: Path, validator: Callable[[pd.DataFrame], bool], max_header_row: int = 10) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
        if validator(df):
            return df
        df = pd.read_csv(path, header=None)
        if validator(df):
            return df
        raise ValueError(f"No se pudo identificar la cabecera en {path.name}")

    last_error: Optional[Exception] = None
    for header_row in range(max_header_row + 1):
        try:
            df = pd.read_excel(path, sheet_name=0, header=header_row)
            if validator(df):
                return df
        except Exception as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    raise ValueError(f"No se pudo identificar la cabecera en {path.name}")


def is_scan_df(df: pd.DataFrame) -> bool:
    cols = [normalize_col_name(c) for c in df.columns]
    p_cols = [c for c in cols if re.fullmatch(r"p\d+", c)]
    return "rut" in cols and len(p_cols) >= HEADER_SCAN_REQUIRED_MIN_P


def is_inscritos_df(df: pd.DataFrame) -> bool:
    cols = {normalize_col_name(c) for c in df.columns}
    has_rut = ("rut" in cols) or any(c.startswith("rut_") for c in cols)
    has_email = ("email" in cols) or ("correo_electronico" in cols)
    return has_rut and has_email


def is_instrumento_df(df: pd.DataFrame) -> bool:
    cols = {normalize_col_name(c) for c in df.columns}
    required = {"pregunta", "respuesta_correcta", "eje", "habilidad", "unidad_tematica"}
    alt = {"pregunta", "respuesta", "eje", "habilidad", "unidad_tematica"}
    return required.issubset(cols) or alt.issubset(cols)


SCAN_ALIASES = {
    "archivo": "archivo",
    "rut": "rut",
    "rut_estado": "rut_estado",
    "observacion": "observacion",
}
INSTRUMENTO_ALIASES = {
    "pregunta": "pregunta",
    "respuesta_correcta": "respuesta_correcta",
    "respuesta": "respuesta_correcta",
    "puntaje": "puntaje",
    "item_id": "item_id",
    "eje": "eje",
    "habilidad": "habilidad",
    "unidad_tematica": "unidad_tematica",
    "nivel_dificultad": "nivel_dificultad",
    "usa_en_puntaje_paes": "usa_en_puntaje_paes",
    "considera_paes": "usa_en_puntaje_paes",
    "incluye_en_puntaje_paes": "usa_en_puntaje_paes",
}


def canonicalize_columns(df: pd.DataFrame, alias_map: dict[str, str], uppercase_questions: bool = False) -> pd.DataFrame:
    rename = {}
    for col in df.columns:
        norm = normalize_col_name(col)
        if norm in alias_map:
            rename[col] = alias_map[norm]
        elif re.fullmatch(r"p\d+", norm):
            rename[col] = norm.upper() if uppercase_questions else norm
    return df.rename(columns=rename).copy()


def coalesce_by_normalized_name(df: pd.DataFrame, candidate_names: list[str]) -> pd.Series:
    norm_candidates = {normalize_col_name(c) for c in candidate_names}
    out = pd.Series([""] * len(df), index=df.index, dtype="object")
    for col in df.columns:
        if normalize_col_name(col) in norm_candidates:
            vals = df[col].fillna("").astype(str).str.strip()
            out = out.where(out.astype(str).str.strip() != "", vals)
    return out


def pick_best_email(df: pd.DataFrame) -> pd.Series:
    candidates: list[pd.Series] = []
    for col in df.columns:
        norm = normalize_col_name(col)
        if norm in {"email", "correo", "correo_electronico"} or norm.startswith("correo_electronico"):
            candidates.append(df[col].fillna("").astype(str).str.strip())

    out = pd.Series([""] * len(df), index=df.index, dtype="object")
    for series in candidates:
        cleaned = series.copy()
        cleaned = cleaned.mask(cleaned.str.lower().isin({"anonymous", "nan", "none", "null"}), "")
        cleaned = cleaned.mask(~cleaned.str.contains("@", regex=False), "")
        out = out.where(out.astype(str).str.strip() != "", cleaned)
    return out


def load_scan_results(path: Path) -> pd.DataFrame:
    df = try_read_with_header(path, is_scan_df)
    df = canonicalize_columns(df, SCAN_ALIASES, uppercase_questions=True)
    q_cols = sorted([c for c in df.columns if re.fullmatch(r"P\d+", str(c))], key=lambda x: int(x[1:]))
    if "rut" not in df.columns:
        raise ValueError("Falta la columna RUT en resultados.xlsx")
    keep_cols = [c for c in ["archivo", "rut", "rut_estado", "observacion"] if c in df.columns] + q_cols
    df = df[keep_cols].copy()
    if "archivo" not in df.columns:
        df["archivo"] = ""
    if "rut_estado" not in df.columns:
        df["rut_estado"] = ""
    if "observacion" not in df.columns:
        df["observacion"] = ""
    for col in q_cols:
        df[col] = (
            df[col]
            .astype(str)
            .str.strip()
            .str.upper()
            .replace({"NAN": "", "NONE": "", "VACIO": "", "VACÍA": "", "VACIA": "", "INV": "", "INVALIDO": ""})
        )
        df[col] = df[col].where(df[col].isin(VALID_ANSWER_CODES | {""}), "N")
        df[col] = df[col].replace({"": "N"})
    df["rut"] = df["rut"].astype(str).str.strip()
    return df


def load_inscritos(path: Path) -> pd.DataFrame:
    raw = try_read_with_header(path, is_inscritos_df)

    out = pd.DataFrame(index=raw.index)
    out["rut"] = coalesce_by_normalized_name(raw, ["rut", "rut_formato123456789"])
    out["nombre_completo"] = coalesce_by_normalized_name(raw, ["nombre_completo"])
    out["email"] = pick_best_email(raw)
    out["curso"] = coalesce_by_normalized_name(raw, ["curso"])
    out["establecimiento"] = coalesce_by_normalized_name(raw, ["establecimiento", "colegio_nombre_completo", "colegio"])
    out["fecha_inscripcion"] = coalesce_by_normalized_name(raw, ["fecha_inscripcion", "hora_de_inicio"])
    out["consentimiento_correo"] = coalesce_by_normalized_name(
        raw,
        [
            "consentimiento_correo",
            "consentimiento",
            "autorizo_a_la_universidad_de_concepcion_a_enviarme_informacion_institucional_invitaciones_a_charlas_y_actividades_academicas_a_traves_de_los_datos_proporcionados",
        ],
    )
    out["telefono"] = coalesce_by_normalized_name(raw, ["telefono", "telefono_opcional_ejemplo_56912345678"])
    out["interes_carreras"] = coalesce_by_normalized_name(
        raw,
        ["interes_carreras", "si_estas_considerando_estudiar_ingenieria_menciona_las_carreras_que_son_de_tu_interes"],
    )

    out["rut"] = out["rut"].astype(str).str.strip()
    out["nombre_completo"] = out["nombre_completo"].astype(str).str.strip()
    out["email"] = out["email"].astype(str).str.strip()
    out["consentimiento_correo"] = out["consentimiento_correo"].map(normalize_yes)

    required = ["rut", "email"]
    missing = [c for c in required if out[c].astype(str).str.strip().eq("").all()]
    if missing:
        raise ValueError(f"Faltan columnas obligatorias o sin datos en inscritos: {missing}")

    return out[
        [
            "rut",
            "nombre_completo",
            "email",
            "curso",
            "establecimiento",
            "fecha_inscripcion",
            "consentimiento_correo",
            "telefono",
            "interes_carreras",
        ]
    ].copy()


def load_instrumento(path: Path) -> pd.DataFrame:
    df = try_read_with_header(path, is_instrumento_df)
    df = canonicalize_columns(df, INSTRUMENTO_ALIASES)
    defaults = {
        "puntaje": 1,
        "item_id": "",
        "nivel_dificultad": "",
        "usa_en_puntaje_paes": "",
    }
    for col, default in defaults.items():
        if col not in df.columns:
            df[col] = default
    required = [
        "pregunta",
        "respuesta_correcta",
        "puntaje",
        "item_id",
        "eje",
        "habilidad",
        "unidad_tematica",
        "nivel_dificultad",
        "usa_en_puntaje_paes",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Faltan columnas en instrumento: {missing}")
    df = df[required].copy()
    df["pregunta"] = pd.to_numeric(df["pregunta"], errors="coerce")
    df = df.dropna(subset=["pregunta"]).copy()
    df["pregunta"] = df["pregunta"].astype(int)
    df["respuesta_correcta"] = df["respuesta_correcta"].astype(str).str.strip().str.upper()
    df["respuesta_correcta"] = df["respuesta_correcta"].where(df["respuesta_correcta"].isin(list("ABCDE")), "A")
    df["puntaje"] = pd.to_numeric(df["puntaje"], errors="coerce").fillna(1.0)
    for col in ["item_id", "eje", "habilidad", "unidad_tematica", "nivel_dificultad"]:
        df[col] = df[col].fillna("").astype(str).str.strip()
    df["usa_en_puntaje_paes"] = df["usa_en_puntaje_paes"].map(normalize_yes)
    if (df["usa_en_puntaje_paes"] == "").all():
        df = df.sort_values("pregunta").reset_index(drop=True)
        df["usa_en_puntaje_paes"] = ["SI" if i < 50 else "NO" for i in range(len(df))]
    return df.sort_values("pregunta").reset_index(drop=True)


def validate_question_columns(scan_df: pd.DataFrame, instrumento_df: pd.DataFrame) -> List[str]:
    expected = [f"P{q}" for q in instrumento_df["pregunta"].tolist()]
    missing = [q for q in expected if q not in scan_df.columns]
    if missing:
        raise ValueError(f"Faltan columnas de preguntas en resultados.xlsx: {missing[:10]}")
    return expected


def compute_item_detail(
    scan_df: pd.DataFrame,
    instrumento_df: pd.DataFrame,
    inscritos_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    question_cols = validate_question_columns(scan_df, instrumento_df)

    scan = scan_df.copy()
    inscritos = inscritos_df.copy()

    scan["rut_normalizado"] = scan["rut"].map(normalize_rut)
    inscritos["rut_normalizado"] = inscritos["rut"].map(normalize_rut)
    inscritos["consentimiento_correo"] = inscritos["consentimiento_correo"].fillna("").astype(str).str.strip().str.upper()

    dup_counts = (
        inscritos["rut_normalizado"]
        .dropna()
        .loc[lambda s: s.astype(str).str.strip() != ""]
        .value_counts()
        .rename_axis("rut_normalizado")
        .reset_index(name="veces_en_inscritos")
    )

    duplicated_ruts = set(
        dup_counts.loc[dup_counts["veces_en_inscritos"] > 1, "rut_normalizado"].tolist()
    )

    duplicados_inscritos_df = inscritos[
        inscritos["rut_normalizado"].isin(duplicated_ruts)
    ].copy()

    if not duplicados_inscritos_df.empty:
        duplicados_inscritos_df = duplicados_inscritos_df.merge(
            dup_counts,
            on="rut_normalizado",
            how="left",
        ).sort_values(["rut_normalizado", "email", "nombre_completo"], kind="stable")
    else:
        duplicados_inscritos_df = pd.DataFrame(
            columns=list(inscritos.columns) + ["rut_normalizado", "veces_en_inscritos"]
        )

    inscritos_validos = inscritos[
        inscritos["rut_normalizado"].notna()
        & (inscritos["rut_normalizado"].astype(str).str.strip() != "")
    ].copy()
    inscritos_validos = inscritos_validos.drop_duplicates(subset=["rut_normalizado"], keep="first")

    merged = scan.merge(
        inscritos_validos,
        on="rut_normalizado",
        how="left",
        suffixes=("_lectura", "_inscrito"),
        indicator=True,
    )

    merged["rut_duplicado_en_inscritos"] = merged["rut_normalizado"].isin(duplicated_ruts)
    merged["encontrado_en_inscritos"] = merged["_merge"].eq("both")
    merged.drop(columns=["_merge"], inplace=True)

    merged["rut_lectura"] = merged["rut_lectura"] if "rut_lectura" in merged.columns else merged["rut"]
    if "rut_inscrito" not in merged.columns:
        merged["rut_inscrito"] = merged["rut"]

    base_cols = [
        "archivo",
        "rut_lectura",
        "rut_estado",
        "observacion",
        "rut_normalizado",
        "rut_inscrito",
        "nombre_completo",
        "email",
        "curso",
        "establecimiento",
        "fecha_inscripcion",
        "consentimiento_correo",
        "telefono",
        "interes_carreras",
        "encontrado_en_inscritos",
        "rut_duplicado_en_inscritos",
    ]
    for col in base_cols:
        if col not in merged.columns:
            merged[col] = ""

    long_df = merged.melt(
        id_vars=base_cols,
        value_vars=question_cols,
        var_name="pregunta_col",
        value_name="respuesta_estudiante",
    )
    long_df["pregunta"] = long_df["pregunta_col"].str.extract(r"(\d+)").astype(int)
    long_df["respuesta_estudiante"] = long_df["respuesta_estudiante"].fillna("").astype(str).str.strip().str.upper()
    long_df["respuesta_estudiante"] = long_df["respuesta_estudiante"].where(
        long_df["respuesta_estudiante"].isin(VALID_ANSWER_CODES),
        "N",
    )

    detail = long_df.merge(instrumento_df, on="pregunta", how="left")

    def classify_response(resp: str, correct: str) -> str:
        if resp == "RD":
            return "RD"
        if resp == "N":
            return "N"
        if resp == correct:
            return "CORRECTA"
        return "INCORRECTA"

    detail["estado_respuesta"] = [
        classify_response(resp, corr)
        for resp, corr in zip(detail["respuesta_estudiante"], detail["respuesta_correcta"], strict=False)
    ]
    detail["es_correcta"] = (detail["estado_respuesta"] == "CORRECTA").astype(int)
    detail["puntaje_logrado"] = detail["puntaje"].where(detail["es_correcta"] == 1, 0)
    detail["considera_en_paes"] = detail["usa_en_puntaje_paes"].fillna("").astype(str).str.upper().eq("SI")
    detail["correcta_paes"] = (detail["es_correcta"].eq(1) & detail["considera_en_paes"]).astype(int)
    detail["nombre_completo"] = detail["nombre_completo"].fillna("").astype(str).str.strip()

    return detail, duplicados_inscritos_df


def paes_from_correct(count: int) -> int:
    count = int(max(0, min(50, count)))
    return PAES_M2_2024_TABLE[count]


def aggregate_student_summary(detail: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    base = (
        detail.groupby("rut_normalizado", dropna=False)
        .agg(
            archivo=("archivo", "first"),
            rut_lectura=("rut_lectura", "first"),
            rut_estado=("rut_estado", "first"),
            observacion=("observacion", "first"),
            rut_inscrito=("rut_inscrito", "first"),
            nombre_completo=("nombre_completo", "first"),
            email=("email", "first"),
            curso=("curso", "first"),
            establecimiento=("establecimiento", "first"),
            fecha_inscripcion=("fecha_inscripcion", "first"),
            consentimiento_correo=("consentimiento_correo", "first"),
            telefono=("telefono", "first"),
            interes_carreras=("interes_carreras", "first"),
            encontrado_en_inscritos=("encontrado_en_inscritos", "max"),
            rut_duplicado_en_inscritos=("rut_duplicado_en_inscritos", "max"),
            total_preguntas=("pregunta", "nunique"),
            puntaje_obtenido=("puntaje_logrado", "sum"),
            puntaje_maximo=("puntaje", "sum"),
            respuestas_correctas=("es_correcta", "sum"),
            correctas_paes=("correcta_paes", "sum"),
            total_preguntas_paes=("considera_en_paes", "sum"),
        )
        .reset_index()
    )

    counts = (
        detail.groupby(["rut_normalizado", "estado_respuesta"], dropna=False)
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    summary = base.merge(counts, on="rut_normalizado", how="left")

    for col in ["INCORRECTA", "N", "RD"]:
        if col not in summary.columns:
            summary[col] = 0

    summary["respuestas_incorrectas"] = summary["INCORRECTA"]
    summary["respuestas_omitidas"] = summary["N"]
    summary["respuestas_dobles"] = summary["RD"]
    summary["logro_global"] = (
        summary["puntaje_obtenido"]
        / summary["puntaje_maximo"].where(summary["puntaje_maximo"] != 0, pd.NA)
    ) * 100
    summary["puntaje_paes"] = summary["correctas_paes"].map(paes_from_correct)
    summary["listo_para_correo"] = (
        summary["encontrado_en_inscritos"]
        & (summary["consentimiento_correo"].fillna("").astype(str).str.upper() == "SI")
        & (summary["rut_estado"].fillna("") == "OK")
    )

    eje = (
        detail.groupby(["rut_normalizado", "eje"], dropna=False)
        .agg(
            puntaje_obtenido=("puntaje_logrado", "sum"),
            puntaje_maximo=("puntaje", "sum"),
        )
        .reset_index()
    )
    eje["logro_eje"] = (
        eje["puntaje_obtenido"]
        / eje["puntaje_maximo"].where(eje["puntaje_maximo"] != 0, pd.NA)
    ) * 100

    habilidad = (
        detail.groupby(["rut_normalizado", "habilidad"], dropna=False)
        .agg(
            puntaje_obtenido=("puntaje_logrado", "sum"),
            puntaje_maximo=("puntaje", "sum"),
        )
        .reset_index()
    )
    habilidad["logro_habilidad"] = (
        habilidad["puntaje_obtenido"]
        / habilidad["puntaje_maximo"].where(habilidad["puntaje_maximo"] != 0, pd.NA)
    ) * 100

    unidad = (
        detail.groupby(["rut_normalizado", "unidad_tematica"], dropna=False)
        .agg(
            puntaje_obtenido=("puntaje_logrado", "sum"),
            puntaje_maximo=("puntaje", "sum"),
        )
        .reset_index()
    )
    unidad["logro_unidad_tematica"] = (
        unidad["puntaje_obtenido"]
        / unidad["puntaje_maximo"].where(unidad["puntaje_maximo"] != 0, pd.NA)
    ) * 100

    eje_wide = eje.pivot(index="rut_normalizado", columns="eje", values="logro_eje")
    eje_wide.columns = [f"eje_{slug_label(c)}_logro" for c in eje_wide.columns]
    eje_wide = eje_wide.reset_index()

    hab_wide = habilidad.pivot(index="rut_normalizado", columns="habilidad", values="logro_habilidad")
    hab_wide.columns = [f"habilidad_{slug_label(c)}_logro" for c in hab_wide.columns]
    hab_wide = hab_wide.reset_index()

    unidad_wide = unidad.pivot(index="rut_normalizado", columns="unidad_tematica", values="logro_unidad_tematica")
    unidad_wide.columns = [f"unidad_{slug_label(c)}_logro" for c in unidad_wide.columns]
    unidad_wide = unidad_wide.reset_index()

    summary = summary.merge(eje_wide, on="rut_normalizado", how="left")
    summary = summary.merge(hab_wide, on="rut_normalizado", how="left")
    summary = summary.merge(unidad_wide, on="rut_normalizado", how="left")
    summary = summary.sort_values(["listo_para_correo", "nombre_completo", "rut_normalizado"], ascending=[False, True, True]).reset_index(drop=True)
    return summary, eje, habilidad, unidad


HTML_STYLE = """
<style>
body { font-family: Arial, sans-serif; color: #111827; margin: 24px; }
h1 { font-size: 28px; text-align: center; margin-bottom: 8px; }
.header-line { text-align: center; margin: 4px 0; font-size: 16px; }
.kpi { text-align: center; font-size: 24px; margin: 16px 0 6px 0; }
.section-title { text-align: center; font-size: 20px; font-weight: 700; margin: 34px 0 12px 0; }
table { width: 100%; border-collapse: collapse; margin-bottom: 22px; }
th { background: #0d4f90; color: white; padding: 10px; font-size: 14px; }
td { background: #edf1f5; padding: 10px; text-align: center; border: 1px solid #d6dee8; }
</style>
"""


def format_pct(value: object) -> str:
    if value is None or pd.isna(value):
        return "0.0%"
    return f"{float(value):.1f}%"


def build_html_file_name(rut_normalizado: object) -> str:
    safe = "" if rut_normalizado is None else str(rut_normalizado).strip()
    safe = safe.replace("\\", "_").replace("/", "_").replace(":", "_").replace("-", "_")
    return f"{safe}.html" if safe else "sin_rut.html"


def _rut_estado_label(value: object) -> str:
    text = "" if value is None else str(value).strip().upper()
    mapping = {
        "OK": "RUT OK",
        "RI": "RUT incompleto",
        "RM": "RUT con múltiples marcas",
        "DV": "DV inconsistente",
        "IMG": "Problema de imagen o marcadores",
    }
    return mapping.get(text, f"RUT: {text}" if text else "RUT no validado")


def build_pending_reasons(summary_df: pd.DataFrame) -> pd.DataFrame:
    out = summary_df.copy()
    reasons = []
    for _, row in out.iterrows():
        row_reasons = []
        rut_estado = str(row.get("rut_estado", "") or "").strip().upper()
        if rut_estado and rut_estado != "OK":
            row_reasons.append(_rut_estado_label(rut_estado))
        if not bool(row.get("encontrado_en_inscritos", False)):
            row_reasons.append("RUT no encontrado en inscritos")
        if str(row.get("consentimiento_correo", "") or "").strip().upper() != "SI":
            row_reasons.append("Sin consentimiento para correo")
        if bool(row.get("rut_duplicado_en_inscritos", False)):
            row_reasons.append("RUT duplicado en inscritos")
        reasons.append("; ".join(dict.fromkeys(row_reasons)))
    out["pendiente_revision_motivo"] = reasons
    out["rut_validacion_detalle"] = out["rut_estado"].map(_rut_estado_label)
    return out


def render_html_report(student: pd.Series, eje_df: pd.DataFrame, hab_df: pd.DataFrame) -> str:
    nombre = html.escape(student.get("nombre_completo") or student.get("rut_normalizado") or "Estudiante")
    rut = html.escape(student.get("rut_normalizado") or "")
    puntaje_paes = int(student.get("puntaje_paes", 100))
    logro_global = format_pct(student.get("logro_global", 0))

    eje_headers = "".join(f"<th>{html.escape(str(v))}</th>" for v in eje_df["eje"].tolist())
    eje_values = "".join(f"<td>{format_pct(v)}</td>" for v in eje_df["logro_eje"].tolist())
    hab_headers = "".join(f"<th>{html.escape(str(v))}</th>" for v in hab_df["habilidad"].tolist())
    hab_values = "".join(f"<td>{format_pct(v)}</td>" for v in hab_df["logro_habilidad"].tolist())

    return f"""<!DOCTYPE html>
<html lang='es'>
<head>
<meta charset='utf-8'>
<title>Informe {nombre}</title>
{HTML_STYLE}
</head>
<body>
  <h1>{nombre}</h1>
  <div class='header-line'>RUT: {rut}</div>
  <div class='kpi'>Puntaje PAES estimado: <strong>{puntaje_paes}</strong></div>
  <div class='header-line'>Logro global: <strong>{logro_global}</strong></div>

  <div class='section-title'>Porcentaje de logro por eje</div>
  <table>
    <tr>{eje_headers}</tr>
    <tr>{eje_values}</tr>
  </table>

  <div class='section-title'>Porcentaje de logro por habilidad</div>
  <table>
    <tr>{hab_headers}</tr>
    <tr>{hab_values}</tr>
  </table>
</body>
</html>
"""


def _pick_first_existing_numeric(df: pd.DataFrame, candidates: list[str], default: float = 0.0) -> pd.Series:
    for col in candidates:
        if col in df.columns:
            return pd.to_numeric(df[col], errors="coerce").fillna(default)
    return pd.Series([default] * len(df), index=df.index, dtype="float64")


def create_mail_merge(summary_df: pd.DataFrame, report_dir: Optional[Path]) -> pd.DataFrame:
    mail = summary_df.copy()
    mail["to_email"] = mail["email"].fillna("").astype(str).str.strip()
    mail["subject"] = "Resultados de ensayo"
    mail["html_file"] = mail["rut_normalizado"].map(build_html_file_name)

    mail["eje_numeros"] = _pick_first_existing_numeric(mail, ["eje_numeros_logro"])
    mail["eje_algebra"] = _pick_first_existing_numeric(mail, ["eje_algebra_logro", "eje_algebra_y_funciones_logro"])
    mail["eje_geometria"] = _pick_first_existing_numeric(mail, ["eje_geometria_logro"])
    mail["eje_datos_azar"] = _pick_first_existing_numeric(
        mail,
        ["eje_datos_y_azar_logro", "eje_probabilidad_y_estadistica_logro"],
    )

    mail["hab_resolver"] = _pick_first_existing_numeric(mail, ["habilidad_resolver_logro"])
    mail["hab_argumentar"] = _pick_first_existing_numeric(mail, ["habilidad_argumentar_logro"])
    mail["hab_modelar"] = _pick_first_existing_numeric(mail, ["habilidad_modelar_logro"])
    mail["hab_representar"] = _pick_first_existing_numeric(mail, ["habilidad_representar_logro"])

    cols = [
        "rut_normalizado",
        "rut_lectura",
        "rut_inscrito",
        "nombre_completo",
        "to_email",
        "subject",
        "listo_para_correo",
        "html_file",
        "puntaje_paes",
        "puntaje_obtenido",
        "logro_global",
        "eje_numeros",
        "eje_algebra",
        "eje_geometria",
        "eje_datos_azar",
        "hab_resolver",
        "hab_argumentar",
        "hab_modelar",
        "hab_representar",
        "respuestas_correctas",
        "total_preguntas",
        "correctas_paes",
        "total_preguntas_paes",
        "rut_estado",
        "encontrado_en_inscritos",
        "rut_duplicado_en_inscritos",
        "consentimiento_correo",
    ]
    return mail[cols].copy()


def autosize_and_style(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    header_fill = PatternFill("solid", fgColor="0D4F90")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        is_mail_merge = ws.title == "mail_merge"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 45)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header_value = ws.cell(1, cell.column).value
                header_norm = normalize_col_name(header_value)
                if isinstance(cell.value, (int, float)) and (
                    "logro" in header_norm
                    or header_norm.startswith("eje_")
                    or header_norm.startswith("hab_")
                    or header_norm.startswith("unidad_")
                ):
                    cell.number_format = '0.0"%"'

        if is_mail_merge:
            if ws.max_row >= 1 and ws.max_column >= 1:
                ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                if "mail_merge_tbl" in ws.tables:
                    del ws.tables["mail_merge_tbl"]
                table = Table(displayName="mail_merge_tbl", ref=ref)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                ws.add_table(table)
        else:
            ws.auto_filter.ref = ws.dimensions

    wb.save(workbook_path)


def build_outputs(
    scan_results_path: Path,
    inscritos_path: Path,
    instrumento_path: Path,
    output_path: Path,
    report_dir: Optional[Path],
) -> None:
    scan_df = load_scan_results(scan_results_path)
    inscritos_df = load_inscritos(inscritos_path)
    instrumento_df = load_instrumento(instrumento_path)

    detail_df, duplicados_inscritos_df = compute_item_detail(scan_df, instrumento_df, inscritos_df)
    summary_df, eje_df, hab_df, unidad_df = aggregate_student_summary(detail_df)
    summary_df = build_pending_reasons(summary_df)

    if report_dir is not None:
        report_dir.mkdir(parents=True, exist_ok=True)
        for _, student in summary_df.iterrows():
            eje_student = eje_df[eje_df["rut_normalizado"] == student["rut_normalizado"]].sort_values("eje")
            hab_student = hab_df[hab_df["rut_normalizado"] == student["rut_normalizado"]].sort_values("habilidad")
            html_content = render_html_report(student, eje_student, hab_student)
            report_path = report_dir / build_html_file_name(student['rut_normalizado'])
            report_path.write_text(html_content, encoding="utf-8")

    pendientes_df = summary_df[
        (~summary_df["listo_para_correo"])
        | (summary_df["rut_estado"].fillna("") != "OK")
        | (~summary_df["encontrado_en_inscritos"])
        | (summary_df["rut_duplicado_en_inscritos"].fillna(False))
    ].copy()

    mail_df = create_mail_merge(summary_df, report_dir)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="resumen_estudiantes")
        eje_df.to_excel(writer, index=False, sheet_name="detalle_por_eje")
        hab_df.to_excel(writer, index=False, sheet_name="detalle_por_habilidad")
        unidad_df.to_excel(writer, index=False, sheet_name="detalle_por_unidad_tematica")
        detail_df.to_excel(writer, index=False, sheet_name="detalle_respuestas")
        pendientes_df.to_excel(writer, index=False, sheet_name="pendientes_revision")
        mail_df.to_excel(writer, index=False, sheet_name="mail_merge")
        duplicados_inscritos_df.to_excel(writer, index=False, sheet_name="duplicados_inscritos")

    autosize_and_style(output_path)

    print(f"Archivo consolidado generado: {output_path}")
    if report_dir is not None:
        print(f"Reportes HTML generados en: {report_dir}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Consolida inputs/resultados.xlsx, inputs/inscritos.xlsx e inputs/instrumento.xlsx."
    )
    parser.add_argument(
        "--scan-results",
        default=DEFAULT_INPUT_DIR / "resultados.xlsx",
        type=Path,
        help="Ruta a inputs/resultados.xlsx",
    )
    parser.add_argument(
        "--inscritos",
        default=DEFAULT_INPUT_DIR / "inscritos.xlsx",
        type=Path,
        help="Ruta a inputs/inscritos.xlsx",
    )
    parser.add_argument(
        "--instrumento",
        default=DEFAULT_INPUT_DIR / "instrumento.xlsx",
        type=Path,
        help="Ruta a inputs/instrumento.xlsx",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_DIR / "resultados_consolidados.xlsx",
        type=Path,
        help="Ruta del Excel consolidado de salida",
    )
    parser.add_argument(
        "--report-dir",
        default=DEFAULT_OUTPUT_DIR / "reportes_html",
        type=Path,
        help="Carpeta de salida para reportes HTML",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build_outputs(
        scan_results_path=args.scan_results,
        inscritos_path=args.inscritos,
        instrumento_path=args.instrumento,
        output_path=args.output,
        report_dir=args.report_dir,
    )


if __name__ == "__main__":
    main()
